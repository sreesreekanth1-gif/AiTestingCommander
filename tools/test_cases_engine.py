"""
╔══════════════════════════════════════════════════════════╗
║   A.N.T. Layer 3 — Test Cases Engine v1                ║
║   Orchestrates: IssuerFetcher → LLMRouter → ExcelBuilder║
╚══════════════════════════════════════════════════════════╝
"""

import json
import os
import sys
import re
from datetime import datetime
import openpyxl

# Ensure tools/ is on the path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from issue_fetcher import fetch_issue
from llm_router import route_to_llm, TOOL_COLUMN_LABEL


# ── Step Normalization & Formatting Helpers ──────────────────────────────────
def normalize_test_steps(raw_steps) -> list:
    """Ensure testSteps is always a list of step dicts regardless of LLM output."""
    if isinstance(raw_steps, list):
        # Validate/fill missing keys
        normalized = []
        for i, s in enumerate(raw_steps):
            if isinstance(s, dict):
                normalized.append({
                    "stepNumber": s.get("stepNumber", i + 1),
                    "action":     s.get("action", s.get("step", "")),
                    "expected":   s.get("expected", s.get("expectedResult", "")),
                    "testData":   s.get("testData", "N/A") or "N/A",
                })
            else:
                normalized.append({"stepNumber": i+1, "action": str(s), "expected": "", "testData": "N/A"})
        return normalized

    # Fallback: flat string — parse into step objects
    raw = str(raw_steps).strip()
    if not raw:
        return [{"stepNumber": 1, "action": "", "expected": "", "testData": "N/A"}]

    lines = [l.strip() for l in re.split(r'\n+', raw) if l.strip()]
    steps = []
    for i, line in enumerate(lines):
        # Strip leading number/bullet
        action = re.sub(r'^\s*(\d+[.)]\s*|[-*•]\s*)', '', line)
        steps.append({"stepNumber": i+1, "action": action, "expected": "", "testData": "N/A"})
    return steps or [{"stepNumber": 1, "action": raw, "expected": "", "testData": "N/A"}]


def derive_summary_fields(tc: dict) -> dict:
    """Compute backward-compatible top-level summaries from steps."""
    steps = tc.get("testSteps", [])

    # Derive expectedResult from steps
    expected_lines = [f"{s['stepNumber']}. {s['expected']}" for s in steps if s.get("expected")]
    tc["expectedResult"] = "\n".join(expected_lines) if expected_lines else ""

    # Derive testData from steps
    data_lines = [f"{s['stepNumber']}. {s['testData']}" for s in steps
                  if s.get("testData") and s["testData"] != "N/A"]
    tc["testData"] = "\n".join(data_lines) if data_lines else "N/A"

    return tc


def format_steps_for_excel(steps: list) -> str:
    """Format structured steps into one Excel cell (numbered, readable)."""
    lines = []
    for s in steps:
        lines.append(f"{s['stepNumber']}. {s['action']}")
        if s.get("expected"):
            lines.append(f"   Expected: {s['expected']}")
        if s.get("testData") and s["testData"] != "N/A":
            lines.append(f"   Test Data: {s['testData']}")
    return "\n".join(lines)

class TestCasesEngine:
    """
    Dedicated engine for generating individual test cases in Excel format.
    Uses 'Test Cases Template.xlsx' as the base.
    """

    def __init__(self, payload_path: str = None):
        self.payload_path = payload_path
        self.config = self._load_payload()
        self.xlsx_template_path = os.path.join(
            os.path.dirname(__file__), "..", "Templates", "Test Cases Template.xlsx"
        )

    def _load_payload(self) -> dict:
        if not self.payload_path:
            return {}
        try:
            with open(self.payload_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            raise RuntimeError(f"TestCasesEngine Failure: Could not load payload. {e}")

    def fetch_issue_context(self) -> str:
        manual = self.config.get("manualRequirements", "").strip()
        if manual:
            print(f"[TC-Engine] Using manual requirements context ({len(manual)} chars).")
            # Build an enriched context if optional fields are present
            enriched = [f"MANUAL REQUIREMENTS:\n{manual}"]
            for field in ["sharedPrerequisites", "businessRules", "widgetsSections", "additionalContext"]:
                val = self.config.get(field, "").strip()
                if val:
                    enriched.append(f"{field.replace('Sections', '').title()}:\n{val}")
            return "\n\n".join(enriched)
        
        return fetch_issue(self.config)

    def generate_test_cases_data(self, context_text: str) -> dict:
        custom = self.config.get("customInstructions", "")
        data = route_to_llm(self.config, context_text, custom_instructions=custom)

        # Normalize and enrich test cases
        test_cases = data.get("testCases", [])
        for tc in test_cases:
            # Normalize testSteps to structured array format
            tc["testSteps"] = normalize_test_steps(tc.get("testSteps", []))
            # Derive summary fields for backward compatibility
            derive_summary_fields(tc)

        data["testCases"] = test_cases
        return data

    def generate_xlsx_file(self, data_schema: dict) -> str:
        """
        Populates the Test Cases template (.xlsx) with dynamic platform headers.
        """
        print("[TC-Engine] Assembling Excel workbook...")
        tool = data_schema.get("selectedTool", self.config.get("selectedTool", "Jira"))
        tool_label = TOOL_COLUMN_LABEL.get(tool, f"{tool} ID")
        test_cases = data_schema.get("testCases", [])
        
        try:
            wb = openpyxl.load_workbook(self.xlsx_template_path)
            sheet = wb.active
        except Exception as e:
            print(f"[TC-Engine] Template not found ({e}), creating blank.")
            wb = openpyxl.Workbook()
            sheet = wb.active
            headers = [
                "Test Case ID", tool_label, "Module/Feature", "Test Case Title",
                "Preconditions", "Test Steps", "Test Data", "Expected Result",
                "Priority", "Test Type"
            ]
            for col, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col, value=header)

        # ── Dynamic Header Mapping ──────────────────────────────────────────
        # Map template header text (lowercased/stripped) to our JSON keys
        HEADER_TO_KEY = {
            "test case id":   "testCaseId",
            "jira id":        "toolTicketId",
            "ado id":         "toolTicketId",
            "ado work item id": "toolTicketId",
            "x-ray id":       "toolTicketId",
            "testrail id":    "toolTicketId",
            "qtest id":       "toolTicketId",
            "module/feature": "module",
            "module":         "module",
            "test case title":"testCaseTitle",
            "preconditions":  "preconditions",
            "test steps":     "testSteps",
            "test data":      "testData",
            "expected result":"expectedResult",
            "priority":       "priority",
            "test type":      "testType"
        }

        from copy import copy

        # Identify which column index maps to which key
        col_map = {}
        template_row_styles = {} # Store styles from row 2 (sample data row)
        
        for col in range(1, sheet.max_column + 1):
            header_val = str(sheet.cell(row=1, column=col).value or "").lower().strip()
            
            # Store the style from the sample row (row 2) for this column
            sample_cell = sheet.cell(row=2, column=col)
            template_row_styles[col] = {
                "font":       copy(sample_cell.font),
                "alignment":  copy(sample_cell.alignment),
                "border":     copy(sample_cell.border),
                "fill":       copy(sample_cell.fill),
                "number_format": sample_cell.number_format,
            }

            # Special case for the dynamic tool label (column 2 usually)
            if "id" in header_val and col == 2:
                col_map["toolTicketId"] = col
                sheet.cell(row=1, column=col, value=tool_label)
            else:
                for h_key, json_key in HEADER_TO_KEY.items():
                    if h_key in header_val:
                        col_map[json_key] = col
                        break

        # ── Write Data starting from row 2 (overwriting sample) ───────────────
        for row_idx, tc in enumerate(test_cases, 2):
            for json_key, col_idx in col_map.items():
                # Format testSteps as readable numbered list in one cell
                if json_key == "testSteps" and isinstance(tc.get("testSteps"), list):
                    val = format_steps_for_excel(tc.get("testSteps", []))
                else:
                    val = tc.get(json_key, "")

                target_cell = sheet.cell(row=row_idx, column=col_idx, value=str(val))

                # Apply the stored template style to the new data cell
                if col_idx in template_row_styles:
                    s = template_row_styles[col_idx]
                    target_cell.font = s["font"]
                    target_cell.alignment = s["alignment"]
                    target_cell.border = s["border"]
                    target_cell.fill = s["fill"]
                    target_cell.number_format = s["number_format"]

        # Output path handling
        if os.name == 'nt':
            tmp_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".tmp"))
        else:
            tmp_dir = "/tmp"
        
        os.makedirs(tmp_dir, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_id = str(self.config.get('issueId', 'output')).replace('/', '_').replace('\\', '_')
        output = os.path.join(tmp_dir, f"TestCases_{output_id}_{ts}.xlsx")
        wb.save(output)
        return output

    def run_pipeline(self) -> tuple:
        """Entry point for Excel + JSON output."""
        print("[TC-Engine] Starting B.L.A.S.T Test Cases Engine Pipeline", file=sys.stderr)
        try:
            context = self.fetch_issue_context()
            print(f"[TC-Engine] Fetched context ({len(context)} chars)", file=sys.stderr)
            test_cases = self.generate_test_cases_data(context)
            print(f"[TC-Engine] Generated {len(test_cases.get('testCases', []))} test cases", file=sys.stderr)
            xlsx_path = self.generate_xlsx_file(test_cases)
            print(f"[TC-Engine] Excel file created at {xlsx_path}", file=sys.stderr)
            return os.path.abspath(xlsx_path), test_cases
        except Exception as e:
            print(f"[TC-Engine] Error: {str(e)}", file=sys.stderr)
            raise
